#!/usr/bin/env python3
"""Build the DM-ready xlsx pack from an enriched wallet CSV.

The enriched CSV must have at least these columns (extras are ignored):
  rank, wallet, volume_usd, swaps, first_trade, last_trade, active_days, platforms
Optional identity columns (blank if absent): ens, x_handle, telegram, farcaster, fc_followers

The script derives per-wallet metrics, bot flags, and profile links, then writes
four tabs: Read Me, All N, DM Shortlist, Platform Stats. Formulas (COUNTIF etc.)
are used for the live stats so the workbook recomputes if rows change - do NOT
hardcode the counts. Run the xlsx skill's recalc.py afterwards; ship only on a
clean recalc.

Usage:
  python3 build_dm_pack.py <enriched_csv> <out_xlsx> \
    --pair "MON/USDC" --chain monad --window "90 days" \
    --query-url "https://dune.com/queries/NNNNNNN" --top-n 1000
"""
import argparse, csv, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
BOT_SWAPS_PER_DAY = 150      # heuristic threshold; document, don't dogmatize
BOT_AVG_TRADE = 100_000


def f(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def i(v, default=0):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def load(path):
    with open(path) as fh:
        rows = list(csv.DictReader(fh))
    if not rows:
        sys.exit("enriched CSV is empty")
    out = []
    for r in rows:
        wallet = r.get("wallet", "").strip()
        if not wallet.startswith("0x"):
            continue
        vol = f(r.get("volume_usd"))
        swaps = i(r.get("swaps"))
        ad = i(r.get("active_days")) or 1
        spd = round(swaps / ad, 2)
        avg = round(vol / swaps, 2) if swaps else 0.0
        out.append({
            "rank": i(r.get("rank")),
            "wallet": wallet,
            "volume_usd": vol,
            "swaps": swaps,
            "first_trade": r.get("first_trade", ""),
            "last_trade": r.get("last_trade", ""),
            "active_days": ad,
            "platforms": r.get("platforms", ""),
            "ens": (r.get("ens") or "").strip(),
            "x_handle": (r.get("x_handle") or "").lstrip("@").strip(),
            "telegram": (r.get("telegram") or "").strip(),
            "farcaster": (r.get("farcaster") or "").strip(),
            "fc_followers": r.get("fc_followers", ""),
            "swaps_per_active_day": spd,
            "avg_trade_usd": avg,
            "likely_bot": "yes" if (spd > BOT_SWAPS_PER_DAY or avg > BOT_AVG_TRADE) else "no",
            "debank_url": f"https://debank.com/profile/{wallet}",
            "arkham_url": f"https://intel.arkm.com/explorer/address/{wallet}",
            "blockscan_chat": f"https://chat.blockscan.com/index?a={wallet}",
        })
    out.sort(key=lambda x: -x["volume_usd"])
    for idx, r in enumerate(out, 1):
        r["rank"] = idx
    return out


def hdr_style(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(vertical="center")


def build(rows, args):
    n = len(rows)
    last = n + 1
    wb = Workbook()

    # ---- Read Me ----
    ws = wb.active
    ws.title = "Read Me"
    intro = [
        (f"Top {n} {args.pair} traders on {args.chain} - DM outreach pack", True, 14),
        (f"Source: Dune query {args.query_url or '(unsaved)'}, dex.trades, "
         f"blockchain={args.chain}, window={args.window}. Ranked by USD volume, descending.", False, 10),
        ("", False, 10),
        ("Column definitions", True, 11),
        ("volume_usd = total USD of the pair's swaps in the window.", False, 10),
        ("platforms = DEX venues that filled the swaps. CAVEAT: aggregator-routed trades "
         "(including Monorail) show under the AMM that filled them, so this is liquidity venue, "
         "not which frontend the user clicked.", False, 10),
        (f"likely_bot = yes when swaps/active-day > {BOT_SWAPS_PER_DAY} or avg trade > "
         f"${BOT_AVG_TRADE:,.0f}. Heuristic - eyeball before excluding anyone.", False, 10),
        ("ens / x_handle / telegram = mainnet ENS reverse record and its public text records "
         "(via ensdata.net). Same EVM address across chains.", False, 10),
        ("farcaster = usually blank: Warpcast public lookup now needs auth. Retry via Neynar with a key.", False, 10),
        ("", False, 10),
        ("Using the DM Shortlist tab", True, 11),
        ("Non-bot wallets only. Identity-resolved rows (ENS or X) sorted to top and shaded. "
         "Fill outreach_status and notes as you work.", False, 10),
        ("No handle? Wallet-native channels: DeBank Hi (debank_url) and Blockscan Chat "
         "(blockscan_chat). [UNVERIFIED] confirm the Blockscan URL for this chain before batch use.", False, 10),
        ("ENS names with no X record are still searchable - paste the ENS name into X search; "
         "people reuse handles.", False, 10),
        ("", False, 10),
        ("Live counts (formulas)", True, 11),
    ]
    for idx, (t, b, sz) in enumerate(intro, start=1):
        ws.cell(row=idx, column=1, value=t).font = Font(name=ARIAL, bold=b, size=sz)
    r0 = len(intro) + 1
    stats = [
        ("Total wallets", f"=COUNTA('All {n}'!B2:B{last})"),
        ("ENS resolved", f"=COUNTIF('All {n}'!I2:I{last},\"?*\")"),
        ("X handles", f"=COUNTIF('All {n}'!J2:J{last},\"?*\")"),
        ("Telegram", f"=COUNTIF('All {n}'!K2:K{last},\"?*\")"),
        ("Likely bots", f"=COUNTIF('All {n}'!P2:P{last},\"yes\")"),
        ("DM shortlist rows", "=COUNTA('DM Shortlist'!B2:B100000)"),
    ]
    for idx, (label, formula) in enumerate(stats):
        ws.cell(row=r0 + idx, column=1, value=label).font = Font(name=ARIAL, size=10)
        c = ws.cell(row=r0 + idx, column=2, value=formula)
        c.font = Font(name=ARIAL, size=10, bold=True)
    ws.column_dimensions["A"].width = 118
    ws.column_dimensions["B"].width = 14

    # ---- All N ----
    ws2 = wb.create_sheet(f"All {n}")
    cols = ["rank", "wallet", "volume_usd", "swaps", "first_trade", "last_trade",
            "active_days", "platforms", "ens", "x_handle", "telegram", "farcaster",
            "swaps_per_active_day", "avg_trade_usd", "debank_url", "likely_bot",
            "arkham_url", "blockscan_chat"]
    ws2.append(cols)
    for r in rows:
        ws2.append([r[c] if not isinstance(r[c], float) else r[c] for c in cols])
    hdr_style(ws2, len(cols))
    for rr in range(2, last + 1):
        ws2.cell(row=rr, column=3).number_format = "$#,##0"
        ws2.cell(row=rr, column=14).number_format = "$#,##0"
        ws2.cell(row=rr, column=13).number_format = "#,##0.0"
        for cc in range(1, len(cols) + 1):
            ws2.cell(row=rr, column=cc).font = Font(name=ARIAL, size=10)
    for idx, w in enumerate([6, 44, 13, 9, 19, 19, 11, 42, 18, 16, 12, 12, 12, 12, 60, 9, 66, 66], 1):
        ws2.column_dimensions[get_column_letter(idx)].width = w
    ws2.freeze_panes = "C2"

    # ---- DM Shortlist ----
    ws3 = wb.create_sheet("DM Shortlist")
    scols = ["rank", "wallet", "ens", "x_handle", "telegram", "volume_usd", "swaps",
             "active_days", "platforms", "avg_trade_usd", "debank_url", "blockscan_chat",
             "outreach_status", "notes"]
    ws3.append(scols)
    nonbot = [r for r in rows if r["likely_bot"] == "no"]
    nonbot.sort(key=lambda r: (0 if (r["ens"] or r["x_handle"]) else 1, -r["volume_usd"]))
    idhits = sum(1 for r in nonbot if r["ens"] or r["x_handle"])
    for r in nonbot:
        ws3.append([r["rank"], r["wallet"], r["ens"], r["x_handle"], r["telegram"],
                    r["volume_usd"], r["swaps"], r["active_days"], r["platforms"],
                    r["avg_trade_usd"], r["debank_url"], r["blockscan_chat"], "", ""])
    hdr_style(ws3, len(scols))
    for rr in range(2, len(nonbot) + 2):
        ws3.cell(row=rr, column=6).number_format = "$#,##0"
        ws3.cell(row=rr, column=10).number_format = "$#,##0"
        for cc in range(1, len(scols) + 1):
            ws3.cell(row=rr, column=cc).font = Font(name=ARIAL, size=10)
        if rr - 2 < idhits:
            for cc in range(1, len(scols) + 1):
                ws3.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor="FFF2CC")
    for idx, w in enumerate([6, 44, 18, 16, 12, 13, 9, 11, 42, 12, 60, 66, 16, 30], 1):
        ws3.column_dimensions[get_column_letter(idx)].width = w
    ws3.freeze_panes = "C2"

    # ---- Platform Stats ----
    ws4 = wb.create_sheet("Platform Stats")
    ws4.append(["platform", "wallets_using_it", "share"])
    hdr_style(ws4, 3)
    seen = {}
    for r in rows:
        for p in filter(None, r["platforms"].split("|")):
            seen[p] = seen.get(p, 0) + 1
    for idx, p in enumerate(sorted(seen, key=seen.get, reverse=True), start=2):
        ws4.cell(row=idx, column=1, value=p).font = Font(name=ARIAL, size=10)
        c1 = ws4.cell(row=idx, column=2, value=f"=COUNTIF('All {n}'!H2:H{last},\"*{p}*\")")
        c1.font = Font(name=ARIAL, size=10)
        c2 = ws4.cell(row=idx, column=3, value=f"=B{idx}/{n}")
        c2.font = Font(name=ARIAL, size=10)
        c2.number_format = "0.0%"
    note_row = len(seen) + 3
    ws4.cell(row=note_row, column=1,
             value="Wallets use multiple venues, so shares exceed 100%. Aggregator-routed "
                   "volume lands on the underlying venue, not the aggregator.").font = \
        Font(name=ARIAL, size=9, italic=True)
    for idx, w in enumerate([18, 16, 12], 1):
        ws4.column_dimensions[get_column_letter(idx)].width = w

    wb.save(args.out_xlsx)
    print(f"saved {args.out_xlsx}: {n} wallets, {len(nonbot)} shortlist rows, {idhits} identity-resolved")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("enriched_csv")
    ap.add_argument("out_xlsx")
    ap.add_argument("--pair", default="TOKEN/USDC")
    ap.add_argument("--chain", default="monad")
    ap.add_argument("--window", default="90 days")
    ap.add_argument("--query-url", default="")
    ap.add_argument("--top-n", type=int, default=1000)
    args = ap.parse_args()
    rows = load(args.enriched_csv)
    build(rows, args)


if __name__ == "__main__":
    main()
